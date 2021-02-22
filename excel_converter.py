from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from datetime import datetime
import json

from prettify import prettify

# Reads configuration file and sets the XML root tag

with open('configuration.json', 'r', encoding='utf-8') as conf:
    data = conf.read()
    data = json.loads(data)

root = ET.Element('comtec')
root.set('version', '2008')

# This will be set on the constructor method or as a function param
# Must match the configuration naming convention
sheet = 'Endereços'

# Loads the workbook and sets the column headers

wb = load_workbook('addresses.xlsx')
ws = wb[sheet]

header = [cell.value for cell in ws[1]]
sheet_data = []

# Separates each row to a list on the sheet_data array

for row in ws.rows: 
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    sheet_data.append(row_data)

# Verifies if the first row is the same as the header variable

assert sheet_data[0] == header
sheet_data.pop(0)

# Translates column header to XML tag name via configuration file

res = map(lambda x: data[sheet]['ws_columns'][x], header)
res = list(res)

# Transforms the row data into tags inside a child tag of root

for row in sheet_data:
    main = ET.SubElement(root, 'address')
    for i, cell in enumerate(row):
        ET.SubElement(main, res[i]).text = str(cell)

# Save prettified version to results directory

now = datetime.now()
now = now.strftime('%Y-%m-%dT%H-%M-%S')

with open(f'results/{now}.xml', 'w', encoding='utf-8') as file:
    file.write(prettify(root))
